"""
management/dispatch_log.py

Generates the car's daily dispatch log ("Йўл варақаси") as an Excel
workbook, based on the driving school's paper booklet template — one
stacked block per requested date, all on a single worksheet (so the whole
run prints as one job instead of the user having to open and print each
date's tab separately), each block an exact copy of the single-page
template (dispatch_log_page_template.xlsx, itself extracted from one page
of the school's original 110-page pre-built booklet, styles and all) with
only the fields that come from our data filled in. A page break is placed
after every block except the last, so each date still lands on its own
printed page.
"""

import copy
import os
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "assets", "dispatch_log_page_template.xlsx")
MAX_DATES = 62  # two months at a time is plenty for one print run

BLOCK_ROWS = 33  # template's page height (rows 1..33 -> A1:AP33)
BLOCK_GAP = 1  # blank row separating one date's block from the next
MAX_COL = 42  # AP — last used column in the template

# Fixed business hours the garage operates on (source: the original
# template's own reference data), not something that varies per car or
# date.
PLANNED_DEPARTURE_HOUR = 7
PLANNED_RETURN_HOUR = 19

UZ_WEEKDAYS = ["Душанба", "Сешанба", "Чоршанба", "Пайшанба", "Жума", "Шанба", "Якшанба"]
UZ_MONTHS = [
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
]


def _format_uz_date(d):
    return f"{UZ_WEEKDAYS[d.weekday()]} «{d.day}» {UZ_MONTHS[d.month - 1]} {d.year} й."


def _instructor_short_name(full_name):
    """First and last word only (drops a middle patronymic, e.g. "Usmonov
    Sherxon Ibragimovich" -> "Usmonov Sherxon"), matching the template's own
    original name-shortening formula — the field isn't wide enough for a
    full three-part name."""
    words = (full_name or "").split()
    if len(words) <= 2:
        return full_name or ""
    return f"{words[0]} {words[1]}"


def _director_line(director_name):
    """"Раҳбари" (branch director) signature-line cell — prints the branch
    director's name ahead of a trailing underscore for their handwritten
    signature. Falls back to the template's original blank signature line
    if the car's instructor has no branch, or the branch has no director
    set."""
    name = (director_name or "").strip()
    return f"Раҳбари {name}{'_' * 20}" if name else "Раҳбари" + "_" * 39


def _fill_page(ws, car, model_name, plate_number, instructor_name, instructor_license, director_name, d):
    # copy_worksheet doesn't carry over print_area (it's a workbook-level
    # defined name, not a plain sheet attribute) — every sheet needs its own.
    ws.print_area = "A1:AP33"
    ws.cell(row=1, column=42).value = None  # unused external "booklet number" field
    ws.cell(row=3, column=15, value=car.id)
    ws.cell(row=6, column=7, value=_format_uz_date(d))
    ws.cell(row=8, column=14, value=model_name)
    ws.cell(row=9, column=12, value=plate_number)
    ws.cell(row=10, column=9, value=instructor_name)
    ws.cell(row=10, column=17, value=instructor_license)
    ws.cell(row=15, column=7, value=PLANNED_DEPARTURE_HOUR)
    ws.cell(row=17, column=7, value=PLANNED_RETURN_HOUR)
    ws.cell(row=33, column=16, value=_director_line(director_name))


def _copy_block(dest_ws, src_ws, row_offset):
    """
    Copies one filled-in template block (rows 1..BLOCK_ROWS of src_ws) into
    dest_ws, shifted down by row_offset rows — cell values, styles, row
    heights and merged ranges all included.
    """
    for r in range(1, BLOCK_ROWS + 1):
        dest_row = r + row_offset
        row_dim = src_ws.row_dimensions.get(r)
        if row_dim and row_dim.height:
            dest_ws.row_dimensions[dest_row].height = row_dim.height
        for c in range(1, MAX_COL + 1):
            src_cell = src_ws.cell(row=r, column=c)
            if src_cell.value is None and not src_cell.has_style:
                continue
            dest_cell = dest_ws.cell(row=dest_row, column=c, value=src_cell.value)
            if src_cell.has_style:
                dest_cell.font = copy.copy(src_cell.font)
                dest_cell.border = copy.copy(src_cell.border)
                dest_cell.fill = copy.copy(src_cell.fill)
                dest_cell.alignment = copy.copy(src_cell.alignment)
                dest_cell.protection = copy.copy(src_cell.protection)
                dest_cell.number_format = src_cell.number_format

    for merged_range in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(
            start_row=merged_range.min_row + row_offset,
            start_column=merged_range.min_col,
            end_row=merged_range.max_row + row_offset,
            end_column=merged_range.max_col,
        )


def generate_dispatch_log_excel(car, dates):
    """
    Returns the dispatch log workbook as bytes: a single worksheet with one
    stacked block per date in `dates` (a list of `date` objects, used in the
    given order), separated by page breaks so each date still prints on its
    own page.
    """
    if not dates:
        raise ValueError("At least one date is required.")
    if len(dates) > MAX_DATES:
        raise ValueError(f"At most {MAX_DATES} dates can be generated at once.")

    model_name, _, plate_number = (car.car_name or "").partition(" ")
    instructor = car.instructor
    instructor_name = _instructor_short_name(instructor.full_name) if instructor else ""
    instructor_license = (
        f"{instructor.license_series or ''}{instructor.license_number or ''}" if instructor else ""
    )
    branch = instructor.branch if instructor else None
    director_name = branch.director_full_name if branch else None

    template_wb = load_workbook(TEMPLATE_PATH)
    template_ws = template_wb.active

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Yo'l varaqalari"
    out_ws.sheet_view.showGridLines = False

    for c in range(1, MAX_COL + 1):
        letter = get_column_letter(c)
        col_dim = template_ws.column_dimensions.get(letter)
        if col_dim and col_dim.width:
            out_ws.column_dimensions[letter].width = col_dim.width

    block_stride = BLOCK_ROWS + BLOCK_GAP
    for i, d in enumerate(dates):
        block_ws = template_wb.copy_worksheet(template_ws)
        _fill_page(block_ws, car, model_name, plate_number, instructor_name, instructor_license, director_name, d)
        row_offset = i * block_stride
        _copy_block(out_ws, block_ws, row_offset)
        del template_wb[block_ws.title]
        if i < len(dates) - 1:
            out_ws.row_breaks.append(Break(id=row_offset + BLOCK_ROWS))

    out_ws.print_area = f"A1:AP{len(dates) * block_stride - BLOCK_GAP}"

    buf = BytesIO()
    out_wb.save(buf)
    return buf.getvalue()
