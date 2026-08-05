"""
management/views.py

All application views for the Driving School Management app live here.
"""

import os
import re
import uuid
from datetime import timedelta
from urllib.parse import quote

from celery.result import AsyncResult
from django.conf import settings
from django.core.exceptions import FieldDoesNotExist
from django.db.models import Q, OuterRef, Subquery, F, Sum, Prefetch, IntegerField
from django.db.models.functions import TruncMonth, Coalesce
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.throttling import ScopedRateThrottle
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView

from management.contract import generate_contract_pdf
from management.dispatch_log import generate_dispatch_log_excel
from management.lesson_book import generate_lesson_book_pdf
from management.models import Branch, Category, User, Enrollment, Payment, Group, LearningPlace, Agent, Holidays, Car, CarWash, DrivingLessons, AutodromeAccessGrant, Notification, TeacherReview, StudentCertificate, Attendance
from management.serializers import (
    BranchSerializer,
    CategorySerializer,
    StudentSerializer,
    UserSerializer,
    StudentCreateSerializer,
    EnrollmentSerializer,
    PaymentSerializer,
    GroupSerializer,
    LearningPlaceSerializer,
    AgentSerializer,
    HolidaysSerializer,
    CarSerializer,
    CarWashSerializer,
    DrivingLessonsSerializer,
    AutodromeAccessGrantSerializer,
    NotificationSerializer,
    TeacherReviewSerializer,
    StudentCertificateSerializer,
    AttendanceSerializer,
)


# ---------------------------------------------------------------------------
# Login / token refresh — tighter throttling than the rest of the API
# ---------------------------------------------------------------------------
# The default per-endpoint throttle (see REST_FRAMEWORK.DEFAULT_THROTTLE_
# RATES in settings.py) is generous, sized for normal app usage — but login
# is exactly the endpoint a credential-stuffing/brute-force script targets,
# so it gets its own much stricter scope regardless of what the general
# anon rate allows.
class ThrottledTokenObtainPairView(TokenObtainPairView):
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "login"


class ThrottledTokenRefreshView(TokenRefreshView):
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "login"


# ---------------------------------------------------------------------------
# Base viewset: soft-delete (sets is_active=False instead of deleting) +
# automatic branch assignment on create
# ---------------------------------------------------------------------------

def _model_has_own_branch_field(model):
    """True if `model` declares its own `branch` ForeignKey (not a reverse relation)."""
    try:
        field = model._meta.get_field("branch")
    except FieldDoesNotExist:
        return False
    return getattr(field, "many_to_one", False)


class SoftDeleteModelViewSet(viewsets.ModelViewSet):
    """
    A ModelViewSet where `destroy` performs a soft-delete:
    sets `is_active = False` on the instance instead of removing it from the DB.

    Also auto-assigns `branch` on create for any model that has that field:
    superusers may pass an explicit `branch` in the payload (this is how the
    navbar branch selector stamps new records), defaulting to their own branch
    if omitted; every other role is always forced to their own assigned
    branch, regardless of what the client sends.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.is_active = False
        instance.save(update_fields=["is_active"])
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_create(self, serializer):
        model = getattr(getattr(serializer, "Meta", None), "model", None)
        if model is None or not _model_has_own_branch_field(model):
            serializer.save()
            return

        user = self.request.user
        is_super = bool(
            user and user.is_authenticated and (user.is_superuser or user.role == User.Role.SUPERUSER)
        )

        if is_super:
            raw_branch = None
            if hasattr(self.request.data, "get"):
                raw_branch = self.request.data.get("branch")
            if raw_branch not in (None, "", "null"):
                try:
                    serializer.save(branch_id=int(raw_branch))
                    return
                except (TypeError, ValueError):
                    pass
            if getattr(user, "branch_id", None):
                serializer.save(branch_id=user.branch_id)
                return
            serializer.save()
            return

        # Non-superusers: always their own branch, client value ignored entirely.
        serializer.save(branch_id=getattr(user, "branch_id", None))


# ---------------------------------------------------------------------------
# Custom standard pagination
# ---------------------------------------------------------------------------

class StandardPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    # High ceiling so the frontend's "Barchasi" (all) pagination option can
    # fetch an entire table in one request for client-side filtering.
    max_page_size = 100000


def annotate_enrollment_paid_amount(queryset):
    """
    Attaches `paid_amount_agg` to each row via a correlated subquery, in the
    same SQL query as the enrollment list itself — EnrollmentSerializer.
    get_paid_amount() picks it up automatically. Without this, serializing
    N enrollments ran N extra `SUM(amount)` queries (one per row, via
    `obj.payments.filter(...).aggregate(...)`); a list of a few hundred/
    thousand enrollments — which is most of this app's list/detail pages —
    turned into a few hundred/thousand extra round trips to the database,
    the single biggest cause of slow page loads (AgentDetailView,
    CertificatesView, GroupDetailView, etc. all render enrollment lists).
    """
    paid_subquery = (
        Payment.objects.filter(enrollment=OuterRef("pk"), status="accepted", is_active=True)
        .values("enrollment")
        .annotate(total=Sum("amount"))
        .values("total")
    )
    return queryset.annotate(
        paid_amount_agg=Coalesce(Subquery(paid_subquery), 0, output_field=IntegerField())
    )


# select_related targets shared by every place that serializes Enrollment
# rows (EnrollmentViewSet directly, and GroupSerializer's nested
# `enrollments` list) — every one of these fields is accessed via `source=`
# on EnrollmentSerializer, so leaving them unjoined means one extra query
# per row per field.
ENROLLMENT_SELECT_RELATED = (
    "student", "category", "group", "instructor", "coordinator", "agent", "learning_place", "branch",
)


def is_admin_or_superuser(user):
    if not user or not user.is_authenticated:
        return False
    return user.is_superuser or user.role in [User.Role.ADMIN, User.Role.SUPERUSER]


def is_superuser(user):
    if not user or not user.is_authenticated:
        return False
    return user.is_superuser or user.role == User.Role.SUPERUSER


def get_scoped_branch_value(request):
    """
    Returns the branch filter value ("id" as str, branch name, a sentinel for
    "no branch assigned", or None for "no restriction") that should be applied
    to this request.

    Only superusers may switch branches via the `?branch=` query param. Every
    other role is locked to their own assigned branch regardless of what the
    client sends.
    """
    user = getattr(request, "user", None)
    is_super = bool(
        user and user.is_authenticated and (user.is_superuser or user.role == User.Role.SUPERUSER)
    )
    if is_super:
        return request.query_params.get("branch")

    if user and user.is_authenticated and user.branch_id:
        return str(user.branch_id)

    # Authenticated but has no branch assigned: only branch-less rows are visible.
    return "__none__"


def filter_by_branch(queryset, request, branch_field="branch"):
    branch = get_scoped_branch_value(request)
    if branch == "__none__":
        kw_null = {f"{branch_field}__isnull": True}
        return queryset.filter(Q(**kw_null))
    if branch:
        if branch.isdigit():
            kw_id = {f"{branch_field}_id": branch}
            kw_null = {f"{branch_field}__isnull": True}
            queryset = queryset.filter(Q(**kw_id) | Q(**kw_null))
        else:
            kw_name = {f"{branch_field}__name__iexact": branch.strip()}
            kw_null = {f"{branch_field}__isnull": True}
            queryset = queryset.filter(Q(**kw_name) | Q(**kw_null))
    return queryset


# ---------------------------------------------------------------------------
# Branch
# ---------------------------------------------------------------------------

class BranchViewSet(SoftDeleteModelViewSet):
    """CRUD for branches / filials."""

    queryset = Branch.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = BranchSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get("search")
        if search:
            qs = qs.filter(name__icontains=search.strip())
        return qs

    # Branches define the whole system's data-scoping boundaries — creating,
    # renaming, or deleting one is superuser-only, same tier as
    # canReturnMoney in the frontend. This viewset previously had no write
    # restriction at all: any authenticated user (including a student
    # account) could create/edit/delete branches via a direct API call.
    def create(self, request, *args, **kwargs):
        if not is_superuser(request.user):
            return Response({"detail": "Filiallarni faqat superuser boshqarishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not is_superuser(request.user):
            return Response({"detail": "Filiallarni faqat superuser boshqarishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_superuser(request.user):
            return Response({"detail": "Filiallarni faqat superuser boshqarishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_superuser(request.user):
            return Response({"detail": "Filiallarni faqat superuser boshqarishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Holidays
# ---------------------------------------------------------------------------

class HolidaysViewSet(SoftDeleteModelViewSet):
    """CRUD for holidays and official days off."""

    queryset = Holidays.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = HolidaysSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get("search")
        if search:
            qs = qs.filter(holiday_name__icontains=search.strip())
        return qs

    def create(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "Bayram yaratish faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    # create() above was the only restriction here — update/destroy were
    # wide open to any authenticated user.
    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Bayramni faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Bayramni faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Bayramni faqat admin va superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Category
# ---------------------------------------------------------------------------

class CategoryViewSet(SoftDeleteModelViewSet):
    """CRUD for driving licence categories."""

    queryset = Category.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = CategorySerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        return filter_by_branch(qs, self.request, "branch")

    def create(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "Kategoriya yaratish faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    # create() above was the only restriction here — update/destroy were
    # wide open to any authenticated user.
    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Kategoriyani faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Kategoriyani faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Kategoriyani faqat admin va superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# User
# ---------------------------------------------------------------------------

class UserViewSet(SoftDeleteModelViewSet):
    """CRUD for system users (staff/instructors/etc.)."""

    queryset = User.objects.filter(is_active=True).order_by("-updated_at", "-date_joined")
    serializer_class = UserSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        # Default (no ?status=) keeps the class queryset — active users only.
        # ?status=inactive / ?status=all opts into seeing soft-deleted users,
        # e.g. for an admin "show deleted" toggle.
        status_param = self.request.query_params.get("status")
        if status_param == "inactive":
            qs = User.objects.filter(is_active=False).order_by("-updated_at", "-date_joined")
        elif status_param == "all":
            qs = User.objects.all().order_by("-updated_at", "-date_joined")
        else:
            qs = super().get_queryset()
        qs = qs.select_related("branch")
        role = self.request.query_params.get("role")
        search = self.request.query_params.get("search")

        if role:
            qs = qs.filter(role=role)
        if search:
            qs = qs.filter(
                Q(full_name__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(phone__icontains=search)
            )
        return filter_by_branch(qs, self.request, "branch")

    @action(detail=False, methods=["get"])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=False, methods=["post"], url_path="change-password")
    def change_password(self, request):
        """Lets any authenticated user set a new password for themself."""
        new_password = request.data.get("new_password")

        if not new_password:
            return Response(
                {"detail": "Yangi parolni kiriting."},
                status=status.HTTP_400_BAD_REQUEST
            )
        if len(new_password) < 4:
            return Response(
                {"detail": "Yangi parol kamida 4 ta belgidan iborat bo'lishi kerak."},
                status=status.HTTP_400_BAD_REQUEST
            )

        request.user.set_password(new_password)
        request.user.save(update_fields=["password"])
        return Response({"detail": "Parol muvaffaqiyatli o'zgartirildi."})

    def create(self, request, *args, **kwargs):
        if not (request.user and (request.user.is_superuser or request.user.role == User.Role.SUPERUSER)):
            return Response(
                {"detail": "Faqat superuser foydalanuvchi yaratishi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not (request.user and (request.user.is_superuser or request.user.role == User.Role.SUPERUSER)):
            return Response(
                {"detail": "Faqat superuser foydalanuvchini tahrirlashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not (request.user and (request.user.is_superuser or request.user.role == User.Role.SUPERUSER)):
            return Response(
                {"detail": "Faqat superuser foydalanuvchini tahrirlashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not (request.user and (request.user.is_superuser or request.user.role == User.Role.SUPERUSER)):
            return Response(
                {"detail": "Faqat superuser foydalanuvchini o'chirishi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Student (User model with role=STUDENT)
# ---------------------------------------------------------------------------

class StudentViewSet(SoftDeleteModelViewSet):
    """CRUD for driving school students (User model with role=student)."""

    queryset = User.objects.filter(role=User.Role.STUDENT, is_active=True).order_by("-updated_at", "-date_joined")
    serializer_class = StudentSerializer
    pagination_class = StandardPagination

    def get_serializer_class(self):
        if self.action == "create":
            return StudentCreateSerializer
        return StudentSerializer

    def create(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "O'quvchini ro'yxatdan o'tkazish faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "O'quvchini tahrirlash faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "O'quvchini tahrirlash faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().partial_update(request, *args, **kwargs)

    # Students without a group have a NULL group_started_at/group_ends_at;
    # nulls_last keeps them at the bottom regardless of sort direction,
    # instead of Postgres's default of sorting NULLs first on DESC.
    GROUP_ORDERING_FIELDS = {
        "group_started_at": F("group_started_at").asc(nulls_last=True),
        "-group_started_at": F("group_started_at").desc(nulls_last=True),
        "group_ends_at": F("group_ends_at").asc(nulls_last=True),
        "-group_ends_at": F("group_ends_at").desc(nulls_last=True),
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        category = self.request.query_params.get("category")
        status_param = self.request.query_params.get("status")
        search = self.request.query_params.get("search")
        phone = self.request.query_params.get("phone")
        jshshr = self.request.query_params.get("jshshr")
        group_name = self.request.query_params.get("group_name")
        has_group = self.request.query_params.get("has_group")
        ordering = self.request.query_params.get("ordering")

        if category:
            if category.isdigit():
                queryset = queryset.filter(enrollments__category_id=category, enrollments__is_active=True)
            else:
                queryset = queryset.filter(enrollments__category__name=category, enrollments__is_active=True)
        if status_param:
            status_map = {
                'Yangi': Enrollment.Status.NEW,
                'Faol': Enrollment.Status.ENROLLED,
                'Tugatgan': Enrollment.Status.FINISHED,
                'Bekor qilingan': Enrollment.Status.CANCELED,
            }
            mapped_status = status_map.get(status_param, status_param)
            # Match only the student's *current* (latest active) enrollment —
            # matching on "any enrollment has this status" would put a student
            # with both an old canceled enrollment and a new active one under
            # both the "Faol" and "Bekor qilingan" tabs at once, disagreeing
            # with the single status StudentSerializer.get_status actually
            # displays for them.
            current_status = Subquery(
                Enrollment.objects.filter(student=OuterRef("pk"), is_active=True)
                .order_by("-created_at", "-id")
                .values("status")[:1]
            )
            queryset = queryset.annotate(current_enrollment_status=current_status).filter(
                current_enrollment_status=mapped_status
            )
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )
        if phone:
            phone_cleaned = re.sub(r"\D", "", phone)
            queryset = queryset.filter(
                Q(phone__icontains=phone_cleaned) |
                Q(phone2__icontains=phone_cleaned)
            )
        if jshshr:
            queryset = queryset.filter(jshshr__icontains=jshshr)
        if group_name:
            queryset = queryset.filter(enrollments__is_active=True, enrollments__group__name__icontains=group_name)
        if has_group is not None:
            has_group_active_group = Q(enrollments__is_active=True, enrollments__group__isnull=False)
            if has_group.lower() in ("1", "true", "yes"):
                queryset = queryset.filter(has_group_active_group)
            else:
                queryset = queryset.exclude(has_group_active_group)

        queryset = filter_by_branch(queryset, self.request, "branch")
        queryset = queryset.distinct()

        if ordering in self.GROUP_ORDERING_FIELDS:
            # Same "current enrollment" definition as get_current_student_enrollment
            # in serializers.py (latest active enrollment) — so the group dates
            # sorted on here always match the ones StudentSerializer displays.
            active_enrollment = Enrollment.objects.filter(
                student=OuterRef("pk"), is_active=True
            ).order_by("-created_at", "-id")
            queryset = queryset.annotate(
                group_started_at=Subquery(active_enrollment.values("group__started_at")[:1]),
                group_ends_at=Subquery(active_enrollment.values("group__ends_at")[:1]),
            ).order_by(self.GROUP_ORDERING_FIELDS[ordering], "-updated_at", "-date_joined")

        # StudentSerializer's 17 SerializerMethodFields all resolve to a
        # single get_current_student_enrollment(student) call each — without
        # this, every one of those independently re-queries the student's
        # enrollments (get_current_student_enrollment now caches its result
        # per student instance, but that cache is only warm if something has
        # already populated it). Prefetching each student's active
        # enrollments here means the very first call finds the answer
        # already in memory instead of hitting the DB, which is what turned
        # a full StudentsView load (up to page_size=100000 rows) into tens
        # of thousands of queries and made the page hang.
        current_enrollments_qs = Enrollment.objects.filter(is_active=True).select_related(
            "category", "instructor", "coordinator", "agent", "group"
        ).order_by("-created_at", "-id")
        queryset = queryset.prefetch_related(
            Prefetch("enrollments", queryset=current_enrollments_qs, to_attr="current_enrollment_list")
        )

        return queryset


# ---------------------------------------------------------------------------
# Enrollment
# ---------------------------------------------------------------------------

class EnrollmentViewSet(SoftDeleteModelViewSet):
    """CRUD for student enrollments."""

    queryset = Enrollment.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = EnrollmentSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        queryset = annotate_enrollment_paid_amount(
            super().get_queryset().select_related(*ENROLLMENT_SELECT_RELATED)
        )
        student = self.request.query_params.get("student")
        category = self.request.query_params.get("category")
        status = self.request.query_params.get("status")
        instructor = self.request.query_params.get("instructor")
        coordinator = self.request.query_params.get("coordinator")
        agent = self.request.query_params.get("agent")
        group = self.request.query_params.get("group")
        learning_place = self.request.query_params.get("learning_place")
        has_certificate = self.request.query_params.get("has_certificate")

        if student:
            queryset = queryset.filter(student_id=student)
        if category:
            queryset = queryset.filter(category_id=category)
        if status:
            queryset = queryset.filter(status=status)
        if instructor:
            queryset = queryset.filter(instructor_id=instructor)
        if coordinator:
            queryset = queryset.filter(coordinator_id=coordinator)
        if agent:
            queryset = queryset.filter(agent_id=agent)
        if group:
            queryset = queryset.filter(group_id=group)
        if learning_place:
            queryset = queryset.filter(learning_place_id=learning_place)
        if has_certificate and has_certificate.lower() in ("1", "true", "yes"):
            queryset = queryset.exclude(student__certificate_number__isnull=True).exclude(student__certificate_number="")
        return filter_by_branch(queryset, self.request, "branch")

    # Enrollment write access wasn't restricted at all before this — since
    # get_queryset only narrows by explicit filter params (never by
    # "belongs to the requesting user"), any authenticated user, including
    # a plain student, could PATCH/DELETE *any* enrollment in their branch
    # (change enrolled_amount/status/instructor/coordinator/agent, or
    # soft-delete it outright) via a direct API call, bypassing every
    # admin-only control the frontend UI enforces. Matches
    # canEditEnrollmentPlacement in the frontend's auth store.
    @staticmethod
    def _can_write(user):
        return is_admin_or_superuser(user) or (user and user.is_authenticated and user.role == User.Role.MECHANIC)

    def create(self, request, *args, **kwargs):
        if not self._can_write(request.user):
            return Response(
                {"detail": "Ro'yxatga olishni faqat admin, superuser yoki mexanik amalga oshirishi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not self._can_write(request.user):
            return Response(
                {"detail": "Ro'yxatga olishni faqat admin, superuser yoki mexanik tahrirlashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not self._can_write(request.user):
            return Response(
                {"detail": "Ro'yxatga olishni faqat admin, superuser yoki mexanik tahrirlashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not self._can_write(request.user):
            return Response(
                {"detail": "Ro'yxatga olishni faqat admin, superuser yoki mexanik o'chirishi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)

    # Shared by export-contract and export-lesson-book: same access rule as
    # the payment/contract-amount info on the student detail page
    # (canSeePaymentInfo in the frontend auth store) — both documents carry
    # enrollment details that go beyond an admin-only editing action.
    @staticmethod
    def _can_view_documents(user, enrollment):
        return bool(
            is_admin_or_superuser(user)
            or (user and user.is_authenticated and user.role in (User.Role.INSTRUCTOR, User.Role.COORDINATOR))
            or (
                user and user.is_authenticated and user.role == User.Role.STUDENT
                and user.id == enrollment.student_id and enrollment.can_view_payments is not False
            )
        )

    @action(detail=True, methods=["get"], url_path="export-contract")
    def export_contract(self, request, pk=None):
        """PDF of this enrollment's "shartnoma" (student agreement)."""
        enrollment = self.get_object()
        if not self._can_view_documents(request.user, enrollment):
            return Response({"detail": "Ruxsat yo'q."}, status=status.HTTP_403_FORBIDDEN)

        pdf_bytes = generate_contract_pdf(enrollment)
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        safe_name = re.sub(r'[\\/*?:"<>|]', "", enrollment.student.full_name or "shartnoma").strip() or "shartnoma"
        filename = f"{safe_name} - shartnoma.pdf"
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return response

    @action(detail=True, methods=["get"], url_path="export-lesson-book")
    def export_lesson_book(self, request, pk=None):
        """PDF of this enrollment's "Amaliy mashq bajarish varaqasi" (practical driving-lesson exercise booklet)."""
        enrollment = self.get_object()
        if not self._can_view_documents(request.user, enrollment):
            return Response({"detail": "Ruxsat yo'q."}, status=status.HTTP_403_FORBIDDEN)

        pdf_bytes = generate_lesson_book_pdf(enrollment)
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        safe_name = re.sub(r'[\\/*?:"<>|]', "", enrollment.student.full_name or "amaliy_mashq").strip() or "amaliy_mashq"
        filename = f"{safe_name} - amaliy mashq daftarchasi.pdf"
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return response


# ---------------------------------------------------------------------------
# Payment
# ---------------------------------------------------------------------------

class PaymentViewSet(SoftDeleteModelViewSet):
    """CRUD for student payments."""

    queryset = Payment.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = PaymentSerializer
    pagination_class = StandardPagination

    # Payment -> Enrollment -> Group is a straight FK chain (no "first of
    # many" ambiguity like Student -> Enrollment), so ordering can reference
    # it directly with no annotate/Subquery needed. nulls_last keeps
    # payments whose enrollment has no group at the bottom either way.
    ORDERING_FIELDS = {
        "group_started_at": F("enrollment__group__started_at").asc(nulls_last=True),
        "-group_started_at": F("enrollment__group__started_at").desc(nulls_last=True),
        "group_ends_at": F("enrollment__group__ends_at").asc(nulls_last=True),
        "-group_ends_at": F("enrollment__group__ends_at").desc(nulls_last=True),
        "created_at": "created_at",
        "-created_at": "-created_at",
        "amount": "amount",
        "-amount": "-amount",
    }

    def get_queryset(self):
        # select_related covers every FK this serializer's `source=` fields
        # traverse (enrollment, enrollment.student/category/group, user,
        # agent, branch, created_by) — without it, each was one extra query
        # per payment row. student_paid_amount_agg does the same for
        # get_student_paid_amount's per-row aggregate (see
        # annotate_enrollment_paid_amount's docstring for the same pattern
        # on enrollments) — pages listing hundreds/thousands of payments
        # (AgentDetailView, the finance views, PaymentsView) were the other
        # major source of slow loads alongside the enrollment N+1.
        student_paid_agg_subquery = (
            Payment.objects.filter(enrollment=OuterRef("enrollment"), status=Payment.Status.ACCEPTED, is_active=True)
            .values("enrollment")
            .annotate(total=Sum("amount"))
            .values("total")
        )
        queryset = super().get_queryset().select_related(
            "enrollment", "enrollment__student", "enrollment__category", "enrollment__group",
            "user", "agent", "branch", "created_by",
        ).annotate(
            student_paid_amount_agg=Coalesce(Subquery(student_paid_agg_subquery), 0, output_field=IntegerField())
        )
        status = self.request.query_params.get("status")
        method = self.request.query_params.get("method")
        category = self.request.query_params.get("category")
        student_name = self.request.query_params.get("student_name")
        agent_name = self.request.query_params.get("agent_name")
        group_name = self.request.query_params.get("group_name")
        jshshr = self.request.query_params.get("jshshr")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        enrollment = self.request.query_params.get("enrollment")
        student = self.request.query_params.get("student")
        agent = self.request.query_params.get("agent")
        user_param = self.request.query_params.get("user")
        user_role = self.request.query_params.get("user_role")
        created_by = self.request.query_params.get("created_by")
        ordering = self.request.query_params.get("ordering")

        if enrollment:
            queryset = queryset.filter(enrollment_id=enrollment)
        if student:
            # bonus_teacher payments are linked to the student's enrollment
            # for reporting (category/group), but the money itself belongs to
            # the teacher — never show it in a student's own payment history.
            queryset = queryset.filter(enrollment__student_id=student).exclude(status=Payment.Status.BONUS_TEACHER)
        if user_param:
            queryset = queryset.filter(user_id=user_param)
        if user_role:
            queryset = queryset.filter(user__role=user_role)
        if created_by:
            queryset = queryset.filter(created_by_id=created_by)
        if agent:
            queryset = queryset.filter(Q(agent_id=agent) | Q(enrollment__agent_id=agent))
        if status:
            if "," in status:
                queryset = queryset.filter(status__in=[s.strip() for s in status.split(",") if s.strip()])
            else:
                queryset = queryset.filter(status=status)
        if method:
            queryset = queryset.filter(method=method)
        if category:
            if category.isdigit():
                queryset = queryset.filter(enrollment__category_id=category)
            else:
                queryset = queryset.filter(enrollment__category__name=category)
        if student_name:
            queryset = queryset.filter(
                Q(enrollment__student__full_name__icontains=student_name) |
                Q(enrollment__category__name__icontains=student_name) |
                Q(agent__full_name__icontains=student_name) |
                Q(user__first_name__icontains=student_name) |
                Q(user__last_name__icontains=student_name) |
                Q(notes__icontains=student_name)
            )
        if agent_name:
            queryset = queryset.filter(
                Q(agent__full_name__icontains=agent_name) |
                Q(enrollment__agent__full_name__icontains=agent_name)
            )
        if group_name:
            queryset = queryset.filter(enrollment__group__name__icontains=group_name)
        if jshshr:
            queryset = queryset.filter(enrollment__student__jshshr__icontains=jshshr)
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=f"{date_to} 23:59:59.999999")

        branch = get_scoped_branch_value(self.request)
        if branch == "__none__":
            queryset = queryset.filter(Q(branch__isnull=True) & Q(enrollment__branch__isnull=True))
        elif branch:
            if branch.isdigit():
                queryset = queryset.filter(Q(branch_id=branch) | Q(enrollment__branch_id=branch) | Q(branch__isnull=True))
            else:
                queryset = queryset.filter(Q(branch__name__iexact=branch.strip()) | Q(enrollment__branch__name__iexact=branch.strip()) | Q(branch__isnull=True))

        # "student_paid_amount" isn't a real column (it's the sum of the
        # linked enrollment's own accepted payments, computed per-row in the
        # serializer) — ordering by it needs the same sum as a correlated
        # subquery annotation instead of a plain field reference.
        if ordering in ("student_paid_amount", "-student_paid_amount"):
            student_paid_subquery = (
                Payment.objects.filter(enrollment=OuterRef("enrollment"), status=Payment.Status.ACCEPTED, is_active=True)
                .values("enrollment")
                .annotate(total=Sum("amount"))
                .values("total")
            )
            queryset = queryset.annotate(student_paid_sort_value=Subquery(student_paid_subquery))
            sort_expr = F("student_paid_sort_value")
            sort_expr = sort_expr.desc(nulls_last=True) if ordering.startswith("-") else sort_expr.asc(nulls_last=True)
            queryset = queryset.order_by(sort_expr, "-updated_at", "-created_at")
        elif ordering in self.ORDERING_FIELDS:
            queryset = queryset.order_by(self.ORDERING_FIELDS[ordering], "-updated_at", "-created_at")

        return queryset

    @action(detail=False, methods=["get"], url_path="monthly-summary")
    def monthly_summary(self, request):
        """
        Server-side month-by-month totals for the finance dashboard's cards.
        The dashboard used to fetch up to 1000 payments and sum them in JS —
        harmless with a handful of hundred payments, but silently wrong (missing
        rows, wrong totals, "no data" for date ranges that actually have data)
        once a branch has more than 1000 accepted payments, which the legacy
        Excel import made routine. Aggregating in the database instead scales
        to any row count and only ever sends back one row per month.
        """
        status_param = request.query_params.get("status", Payment.Status.ACCEPTED)
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        qs = Payment.objects.filter(is_active=True, status=status_param)
        branch = get_scoped_branch_value(request)
        if branch == "__none__":
            qs = qs.filter(Q(branch__isnull=True) & Q(enrollment__branch__isnull=True))
        elif branch:
            if branch.isdigit():
                qs = qs.filter(Q(branch_id=branch) | Q(enrollment__branch_id=branch) | Q(branch__isnull=True))
            else:
                qs = qs.filter(Q(branch__name__iexact=branch.strip()) | Q(enrollment__branch__name__iexact=branch.strip()) | Q(branch__isnull=True))
        if date_from:
            qs = qs.filter(created_at__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__lte=f"{date_to} 23:59:59.999999")

        rows = (
            qs.annotate(month=TruncMonth("created_at"))
            .values("month", "method")
            .annotate(total=Sum("amount"))
            .order_by("month")
        )

        buckets = {}
        method_keys = {m for m, _ in Payment.Method.choices}
        for row in rows:
            key = row["month"].strftime("%Y-%m")
            bucket = buckets.setdefault(
                key, {"key": key, "cash": 0, "card": 0, "transfer": 0, "qr_code": 0, "click": 0, "total": 0}
            )
            amount = row["total"] or 0
            if row["method"] in method_keys:
                bucket[row["method"]] = bucket.get(row["method"], 0) + amount
            bucket["total"] += amount

        return Response(sorted(buckets.values(), key=lambda b: b["key"], reverse=True))

    def create(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "To'lovni qabul qilish faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        # Returning money is a superuser-only operation — admins may record
        # every other kind of payment but never hand money back.
        if request.data.get("status") == Payment.Status.RETURNED and not is_superuser(request.user):
            return Response(
                {"detail": "Pulni qaytarish faqat superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        # Always the actual logged-in admin/superuser recording this payment —
        # never taken from client input, since "user" already carries a
        # different, status-dependent meaning per row.
        user = self.request.user
        enrollment = serializer.validated_data.get("enrollment")
        if enrollment is not None:
            # A payment tied to an enrollment always belongs to that
            # enrollment's own branch — derived, never left ambiguous
            # (null) or stamped from whichever branch happened to be
            # selected in the admin's UI at the moment of recording it,
            # which could silently mismatch the student's actual branch
            # and make branch-scoped income totals inconsistent.
            serializer.save(created_by=user, branch=enrollment.branch)
            return

        # No enrollment (e.g. a teacher's salary payment) — same branch
        # auto-assignment every other model gets via SoftDeleteModelViewSet.
        is_super = bool(
            user and user.is_authenticated and (user.is_superuser or user.role == User.Role.SUPERUSER)
        )
        if is_super:
            raw_branch = None
            if hasattr(self.request.data, "get"):
                raw_branch = self.request.data.get("branch")
            if raw_branch not in (None, "", "null"):
                try:
                    serializer.save(created_by=user, branch_id=int(raw_branch))
                    return
                except (TypeError, ValueError):
                    pass
            if getattr(user, "branch_id", None):
                serializer.save(created_by=user, branch_id=user.branch_id)
                return
            serializer.save(created_by=user)
            return
        serializer.save(created_by=user, branch_id=getattr(user, "branch_id", None))

    def update(self, request, *args, **kwargs):
        if not (request.user and (request.user.is_superuser or request.user.role == User.Role.SUPERUSER)):
            return Response(
                {"detail": "Faqat superuser to'lov ma'lumotlarini tahrirlashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not (request.user and (request.user.is_superuser or request.user.role == User.Role.SUPERUSER)):
            return Response(
                {"detail": "Faqat superuser to'lov ma'lumotlarini tahrirlashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not (request.user and (request.user.is_superuser or request.user.role == User.Role.SUPERUSER)):
            return Response(
                {"detail": "Faqat superuser to'lov ma'lumotlarini tahrirlashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Group
# ---------------------------------------------------------------------------

class GroupViewSet(SoftDeleteModelViewSet):
    """CRUD for groups of students."""

    queryset = Group.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = GroupSerializer
    pagination_class = StandardPagination

    # Groups without a start/end date sort to the bottom regardless of
    # direction, instead of Postgres's default of NULLs first on DESC.
    ORDERING_FIELDS = {
        "started_at": F("started_at").asc(nulls_last=True),
        "-started_at": F("started_at").desc(nulls_last=True),
        "ends_at": F("ends_at").asc(nulls_last=True),
        "-ends_at": F("ends_at").desc(nulls_last=True),
    }

    def get_queryset(self):
        qs = super().get_queryset().select_related("category", "branch")
        # GroupSerializer nests the group's full enrollment list
        # (EnrollmentSerializer(many=True)) — without this prefetch, that
        # nested list re-triggers the exact same N+1 pattern per group (see
        # annotate_enrollment_paid_amount's docstring), which is what made
        # GroupDetailView slow to load. Scoped to `retrieve` calls that
        # explicitly ask for it via ?with_enrollments=1 — GroupDetailView is
        # the only caller that needs the roster; every other single-group
        # fetch across the app (e.g. StudentDetailView, just to read
        # started_at/ends_at by id) was paying to prefetch and serialize a
        # roster it never reads, same waste as the list case below.
        if self.action == "retrieve" and self.request.query_params.get("with_enrollments"):
            enrollments_qs = annotate_enrollment_paid_amount(
                Enrollment.objects.filter(is_active=True).select_related(*ENROLLMENT_SELECT_RELATED)
            )
            qs = qs.prefetch_related(Prefetch("enrollments", queryset=enrollments_qs))
        status_param = self.request.query_params.get("status")
        category = self.request.query_params.get("category")
        name = self.request.query_params.get("name")
        started_at_from = self.request.query_params.get("started_at_from")
        started_at_to = self.request.query_params.get("started_at_to")
        ends_at_from = self.request.query_params.get("ends_at_from")
        ends_at_to = self.request.query_params.get("ends_at_to")
        ordering = self.request.query_params.get("ordering")

        if status_param:
            qs = qs.filter(status=status_param)
        if category:
            if category.isdigit():
                qs = qs.filter(category_id=category)
            else:
                qs = qs.filter(category__name=category)
        if name:
            qs = qs.filter(name__icontains=name)
        if started_at_from:
            qs = qs.filter(started_at__gte=started_at_from)
        if started_at_to:
            qs = qs.filter(started_at__lte=started_at_to)
        if ends_at_from:
            qs = qs.filter(ends_at__gte=ends_at_from)
        if ends_at_to:
            qs = qs.filter(ends_at__lte=ends_at_to)

        qs = filter_by_branch(qs, self.request, "branch")

        if ordering in self.ORDERING_FIELDS:
            qs = qs.order_by(self.ORDERING_FIELDS[ordering], "-updated_at", "-created_at")

        return qs

    def create(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "Guruh yaratish faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "Guruhni tahrirlash faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "Guruhni tahrirlash faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().partial_update(request, *args, **kwargs)

    @action(detail=True, methods=["get"], url_path="export-payments")
    def export_payments(self, request, pk=None):
        """One-sheet Excel export of this group's "A'zo o'quvchilar" table —
        same columns shown on the group detail page, for offline sharing."""
        if getattr(request.user, "role", None) == User.Role.MECHANIC:
            return Response({"detail": "Ruxsat yo'q."}, status=status.HTTP_403_FORBIDDEN)

        group = self.get_object()
        enrollments = Enrollment.objects.filter(group=group, is_active=True).select_related(
            "student", "learning_place", "instructor", "coordinator", "agent"
        ).order_by("student__full_name")

        weekday_names = ["Dush", "Sesh", "Chor", "Pay", "Juma", "Shan"]
        status_names = {"new": "Yangi", "enrolled": "Faol", "finished": "Tugatgan", "canceled": "Bekor qilingan"}

        def full_name(user):
            if not user:
                return ""
            name = f"{user.first_name} {user.last_name}".strip()
            return name or user.phone

        wb = Workbook()
        ws = wb.active
        ws.title = "To'lovlar"

        # Student contact/ID and group-date columns are Excel-only — not
        # shown in the on-screen "A'zo o'quvchilar" table (which already has
        # its own group start/end columns for the group as a whole; these
        # are duplicated here per-row purely for offline record-keeping).
        headers = [
            "O'quvchi F.I.SH.", "Telefon 1", "Telefon 2", "JSHSHR", "Pasport",
            "Guruh nomi", "Guruh boshlanishi", "Guruh tugashi",
            "Holati", "O'quv joyi", "Dars vaqti", "Dars kunlari",
            "O'qituvchi", "Instruktor", "Agent",
            "Shartnoma summasi", "To'langan", "Qoldiq",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for e in enrollments:
            paid = e.payments.filter(status="accepted", is_active=True).aggregate(total=Sum("amount"))["total"] or 0
            enrolled_amount = 0 if e.enrolled_free else (e.enrolled_amount or 0)
            remaining = 0 if (e.enrolled_free or e.excel_imported) else enrolled_amount - paid
            learning_days = ", ".join(
                weekday_names[d] for d in (e.learning_days or []) if 0 <= d < len(weekday_names)
            )
            passport = ""
            if e.student and (e.student.passport_serie or e.student.passport_number):
                passport = f"{e.student.passport_serie or ''} {e.student.passport_number or ''}".strip()
            ws.append([
                e.student.full_name if e.student else "",
                e.student.phone if e.student else "",
                e.student.phone2 if e.student else "",
                e.student.jshshr if e.student else "",
                passport,
                group.name or "",
                group.started_at.isoformat() if group.started_at else "",
                group.ends_at.isoformat() if group.ends_at else "",
                status_names.get(e.status, e.status),
                e.learning_place.place_name if e.learning_place else "",
                e.learning_time or "",
                learning_days,
                full_name(e.coordinator),
                full_name(e.instructor),
                e.agent.full_name if e.agent else "",
                enrolled_amount,
                0 if e.enrolled_free else paid,
                remaining,
            ])

        for i, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 2)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        safe_name = re.sub(r'[\\/*?:"<>|]', "", group.name or "guruh").strip() or "guruh"
        filename = f"{safe_name} to'lovlar.xlsx"
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        wb.save(response)
        return response


# ---------------------------------------------------------------------------
# LearningPlace
# ---------------------------------------------------------------------------

class LearningPlaceViewSet(SoftDeleteModelViewSet):
    """CRUD for physical / online learning places."""

    queryset = LearningPlace.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = LearningPlaceSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        return filter_by_branch(qs, self.request, "branch")

    def create(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "O'quv joyi yaratish faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    # create() above was the only restriction here — update/destroy were
    # wide open to any authenticated user.
    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "O'quv joyini faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "O'quv joyini faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "O'quv joyini faqat admin va superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Agent
# ---------------------------------------------------------------------------

class AgentViewSet(SoftDeleteModelViewSet):
    """CRUD for Agents (Student recruiters / referrals)."""

    queryset = Agent.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = AgentSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset().select_related("branch", "user")
        search = self.request.query_params.get("search", None)
        if search:
            search_cleaned = search.strip().lower()
            qs = qs.filter(
                Q(full_name__icontains=search_cleaned) |
                Q(phone__icontains=search_cleaned) |
                Q(phone2__icontains=search_cleaned)
            )
        name = self.request.query_params.get("name", None)
        if name:
            qs = qs.filter(full_name__icontains=name.strip())
        phone = self.request.query_params.get("phone", None)
        if phone:
            phone_cleaned = re.sub(r"\D", "", phone)
            qs = qs.filter(Q(phone__icontains=phone_cleaned) | Q(phone2__icontains=phone_cleaned))
        return filter_by_branch(qs, self.request, "branch")

    def create(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "Agent yaratish faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    # create() above was the only restriction here — update/destroy were
    # wide open to any authenticated user.
    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Agentni faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Agentni faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Agentni faqat admin va superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Car
# ---------------------------------------------------------------------------

class CarViewSet(SoftDeleteModelViewSet):
    """CRUD for driving school vehicles."""

    queryset = Car.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = CarSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        queryset = super().get_queryset()
        car_name = self.request.query_params.get("car_name") or self.request.query_params.get("search")
        status_param = self.request.query_params.get("status")
        instructor_param = self.request.query_params.get("instructor")
        instructor_name = self.request.query_params.get("instructor_name")
        if car_name:
            queryset = queryset.filter(car_name__icontains=car_name)
        if status_param:
            queryset = queryset.filter(status=status_param)
        if instructor_param:
            queryset = queryset.filter(instructor_id=instructor_param)
        if instructor_name:
            queryset = queryset.filter(instructor__full_name__icontains=instructor_name)
        return queryset

    # Matches canEditCars in the frontend's auth store (admin/superuser/
    # mechanic) — previously unrestricted, so any authenticated user could
    # create/edit/delete vehicle records via a direct API call. Washing a
    # car has its own narrower check in mark_washed below (also open to the
    # car's assigned instructor), left untouched.
    @staticmethod
    def _can_write(user):
        return is_admin_or_superuser(user) or (user and user.is_authenticated and user.role == User.Role.MECHANIC)

    def create(self, request, *args, **kwargs):
        if not self._can_write(request.user):
            return Response({"detail": "Avtomobillarni faqat admin, superuser yoki mexanik boshqarishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not self._can_write(request.user):
            return Response({"detail": "Avtomobillarni faqat admin, superuser yoki mexanik boshqarishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not self._can_write(request.user):
            return Response({"detail": "Avtomobillarni faqat admin, superuser yoki mexanik boshqarishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not self._can_write(request.user):
            return Response({"detail": "Avtomobillarni faqat admin, superuser yoki mexanik boshqarishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["post"])
    def mark_washed(self, request, pk=None):
        """
        Marks this car as washed. The car's currently assigned instructor,
        any mechanic, or an admin/superuser may do this. `washed_at` is
        always the server's current time — any date/time sent by the client
        is ignored, so a wash can never be backdated.
        """
        car = self.get_object()
        user = request.user

        is_assigned_instructor = bool(
            user and user.is_authenticated and car.instructor_id and user.id == car.instructor_id
        )
        is_mechanic = bool(user and user.is_authenticated and user.role == User.Role.MECHANIC)
        if not (is_assigned_instructor or is_mechanic or is_admin_or_superuser(user)):
            return Response(
                {"detail": "Faqat ushbu avtomobilga biriktirilgan instruktor, mexanik yoki admin uni yuvilgan deb belgilashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )

        wash = CarWash.objects.create(car=car, instructor_id=car.instructor_id, washed_at=timezone.now())
        car.last_washed_at = wash.washed_at
        car.save(update_fields=["last_washed_at", "updated_at"])

        return Response(CarWashSerializer(wash).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"], url_path="export-dispatch-log")
    def export_dispatch_log(self, request, pk=None):
        """
        Excel dispatch log ("Йўл варақаси"), one printable page per date in
        [date_from, date_to] (inclusive). Same access rule as mark_washed —
        the car's assigned instructor, any mechanic, or an admin/superuser.
        """
        car = self.get_object()
        user = request.user

        is_assigned_instructor = bool(
            user and user.is_authenticated and car.instructor_id and user.id == car.instructor_id
        )
        is_mechanic = bool(user and user.is_authenticated and user.role == User.Role.MECHANIC)
        if not (is_assigned_instructor or is_mechanic or is_admin_or_superuser(user)):
            return Response(
                {"detail": "Faqat ushbu avtomobilga biriktirilgan instruktor, mexanik yoki admin yo'l varaqasini yuklab olishi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )

        date_from = parse_date(request.query_params.get("date_from") or "")
        date_to = parse_date(request.query_params.get("date_to") or "")
        if not date_from or not date_to:
            return Response({"detail": "date_from va date_to sanalari kerak."}, status=status.HTTP_400_BAD_REQUEST)
        if date_to < date_from:
            return Response({"detail": "date_to date_from dan oldin bo'lishi mumkin emas."}, status=status.HTTP_400_BAD_REQUEST)

        dates = []
        cursor = date_from
        while cursor <= date_to:
            dates.append(cursor)
            cursor += timedelta(days=1)

        try:
            excel_bytes = generate_dispatch_log_excel(car, dates)
        except ValueError as err:
            return Response({"detail": str(err)}, status=status.HTTP_400_BAD_REQUEST)

        response = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        safe_name = re.sub(r'[\\/*?:"<>|]', "", car.car_name or "yol varaqasi").strip() or "yol varaqasi"
        filename = f"{safe_name} - yo'l varaqasi.xlsx"
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return response


# ---------------------------------------------------------------------------
# DrivingLessons
# ---------------------------------------------------------------------------

class DrivingLessonsViewSet(SoftDeleteModelViewSet):
    """CRUD for driving lesson confirmations."""

    queryset = DrivingLessons.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = DrivingLessonsSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        student = self.request.query_params.get("student")
        instructor = self.request.query_params.get("instructor")
        car = self.request.query_params.get("car")
        lesson_type = self.request.query_params.get("lesson_type")
        if student:
            qs = qs.filter(student_id=student)
        if instructor:
            qs = qs.filter(instructor_id=instructor)
        if car:
            qs = qs.filter(car_id=car)
        if lesson_type:
            qs = qs.filter(lesson_type=lesson_type)
        return filter_by_branch(qs, self.request, "branch")

    # Matches canAddDrivingLesson in the frontend's auth store: a student
    # may confirm a lesson for *themselves* only, admin/superuser for
    # anyone. This viewset previously had no restriction of any kind — any
    # authenticated user could POST a lesson (or autodrome hour) for *any*
    # student's id, fabricating progress/hours, or PATCH/DELETE any
    # existing confirmation (erasing or inflating another student's
    # record) with a direct API call.
    def create(self, request, *args, **kwargs):
        user = request.user
        if is_admin_or_superuser(user):
            return super().create(request, *args, **kwargs)
        if user and user.is_authenticated and user.role == User.Role.STUDENT:
            if str(request.data.get("student")) != str(user.id):
                return Response(
                    {"detail": "Faqat o'zingiz uchun dars tasdiqlashingiz mumkin."},
                    status=status.HTTP_403_FORBIDDEN
                )
            return super().create(request, *args, **kwargs)
        return Response(
            {"detail": "Amaliy dars tasdiqlashni faqat o'quvchi (o'zi uchun) yoki admin/superuser amalga oshirishi mumkin."},
            status=status.HTTP_403_FORBIDDEN
        )

    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Tasdiqlangan darsni faqat admin yoki superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Tasdiqlangan darsni faqat admin yoki superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Tasdiqlangan darsni faqat admin yoki superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# AutodromeAccessGrant
# ---------------------------------------------------------------------------

class AutodromeAccessGrantViewSet(SoftDeleteModelViewSet):
    """Extra avtodrom visit allowances granted by admins/superusers."""

    queryset = AutodromeAccessGrant.objects.filter(is_active=True).order_by("-created_at")
    serializer_class = AutodromeAccessGrantSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        student = self.request.query_params.get("student")
        if student:
            qs = qs.filter(student_id=student)
        return filter_by_branch(qs, self.request, "branch")

    def create(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response(
                {"detail": "Avtodrom uchun qo'shimcha ruxsat berish faqat admin va superuser uchun ruxsat etilgan."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        user = self.request.user
        serializer.save(granted_by=user, branch_id=getattr(user, "branch_id", None))

    # create() above was the only restriction here — update/destroy were
    # wide open to any authenticated user (a student could grant/revoke
    # their own extra avtodrom hours).
    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Ruxsatni faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Ruxsatni faqat admin va superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Ruxsatni faqat admin va superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Notification
# ---------------------------------------------------------------------------

class NotificationViewSet(SoftDeleteModelViewSet):
    """CRUD & Actions for notifications."""

    queryset = Notification.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = NotificationSerializer
    pagination_class = StandardPagination

    # Every real notification in this app is created server-side (by other
    # viewsets' business logic, e.g. TeacherReviewViewSet.perform_create),
    # never posted directly by a client — but this viewset had no create
    # restriction at all, so any authenticated user could POST an arbitrary
    # notification (fake title/status/target, addressed to any user id) —
    # a spam/social-engineering vector. mark_as_read/mark_all_read below
    # are untouched (already properly scoped to the requester's own rows
    # via get_queryset).
    def create(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Bildirishnoma yaratish faqat admin yoki superuser uchun ruxsat etilgan."}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Bildirishnomani faqat admin yoki superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Bildirishnomani faqat admin yoki superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Bildirishnomani faqat admin yoki superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    def get_queryset(self):
        qs = super().get_queryset()
        status_param = self.request.query_params.get("status")
        is_read_param = self.request.query_params.get("is_read")

        if status_param:
            qs = qs.filter(status=status_param)
        if is_read_param is not None:
            is_read_bool = is_read_param.lower() in ["true", "1"]
            qs = qs.filter(is_read=is_read_bool)

        # Every notification is scoped to whoever is asking: admins/superusers
        # see their own plus any "for admins" broadcast (user is null) rows —
        # e.g. new review or certificate-upload alerts — while everyone else
        # only ever sees notifications addressed directly to them.
        user = self.request.user
        if user and user.is_authenticated:
            if is_admin_or_superuser(user):
                qs = qs.filter(Q(user_id=user.id) | Q(user__isnull=True))
            else:
                qs = qs.filter(user_id=user.id)
        else:
            qs = qs.none()

        return filter_by_branch(qs, self.request, "branch")

    @action(detail=False, methods=["get"])
    def unread_count(self, request):
        """Returns total unread count for notifications."""
        user = request.user if request.user and request.user.is_authenticated else None
        qs = Notification.objects.filter(is_active=True, is_read=False)
        if user:
            if is_admin_or_superuser(user):
                qs = qs.filter(Q(user=user) | Q(user__isnull=True))
            else:
                qs = qs.filter(user=user)
        else:
            qs = qs.none()
        qs = filter_by_branch(qs, request, "branch")
        count = qs.count()
        return Response({"unread_count": count})

    @action(detail=True, methods=["post"])
    def mark_as_read(self, request, pk=None):
        """Mark single notification as read."""
        notif = self.get_object()
        notif.is_read = True
        notif.save(update_fields=["is_read", "updated_at"])
        return Response({"status": "marked as read", "is_read": True})

    @action(detail=False, methods=["post"])
    def mark_all_read(self, request):
        """Mark all of the requesting user's unread notifications as read."""
        user = request.user if request.user and request.user.is_authenticated else None
        qs = Notification.objects.filter(is_active=True, is_read=False)
        if user:
            if is_admin_or_superuser(user):
                qs = qs.filter(Q(user=user) | Q(user__isnull=True))
            else:
                qs = qs.filter(user=user)
        else:
            qs = qs.none()
        qs.update(is_read=True)
        return Response({"status": "all marked as read"})


# ---------------------------------------------------------------------------
# TeacherReview
# ---------------------------------------------------------------------------

class TeacherReviewViewSet(SoftDeleteModelViewSet):
    """Reviews students leave for their coordinator/instructor."""

    queryset = TeacherReview.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = TeacherReviewSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        student = self.request.query_params.get("student")
        teacher = self.request.query_params.get("teacher")
        if student:
            qs = qs.filter(student_id=student)
        if teacher:
            qs = qs.filter(teacher_id=teacher)
        return filter_by_branch(qs, self.request, "branch")

    def create(self, request, *args, **kwargs):
        user = request.user
        if not (user and user.is_authenticated):
            return Response(
                {"detail": "Avtorizatsiyadan o'ting."},
                status=status.HTTP_401_UNAUTHORIZED
            )
        teacher_id = request.data.get("teacher")
        if teacher_id and str(teacher_id) == str(user.id):
            return Response(
                {"detail": "O'zingizga sharh qoldira olmaysiz."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        user = self.request.user
        review = serializer.save(student=user, branch_id=getattr(user, "branch_id", None))

        # Notify admins/superusers only — this Notification has no `user` set,
        # so NotificationViewSet's queryset only surfaces it to admin/superuser.
        teacher_label = review.teacher.full_name or review.teacher.phone
        student_label = review.student.full_name or review.student.phone
        Notification.objects.create(
            title=f"Yangi sharh: {student_label} → {teacher_label}",
            note=review.comment or f"Baho: {review.rating}/5",
            status=Notification.Status.REVIEW,
            target_id=review.teacher_id,
            user=None,
            branch=None,
        )

    # create() above validates who may post a review; update/destroy had no
    # check at all — any authenticated user could edit or delete *any*
    # review (not just their own), including the teacher being reviewed
    # editing away a bad one. No frontend flow edits an existing review
    # (it's a rate-once action), so this is moderation-only.
    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Sharhni faqat admin yoki superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Sharhni faqat admin yoki superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Sharhni faqat admin yoki superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# StudentCertificate
# ---------------------------------------------------------------------------

class StudentCertificateViewSet(SoftDeleteModelViewSet):
    """Exam-pass certificates a teacher (coordinator) uploads for a student."""

    queryset = StudentCertificate.objects.filter(is_active=True).order_by("-updated_at", "-created_at")
    serializer_class = StudentCertificateSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        student = self.request.query_params.get("student")
        coordinator = self.request.query_params.get("coordinator")
        if student:
            qs = qs.filter(student_id=student)
        if coordinator:
            qs = qs.filter(coordinator_id=coordinator)
        return filter_by_branch(qs, self.request, "branch")

    def create(self, request, *args, **kwargs):
        user = request.user
        if not (user and user.is_authenticated and (
            user.role == User.Role.COORDINATOR or user.is_superuser or user.role == User.Role.SUPERUSER
        )):
            return Response(
                {"detail": "Faqat o'qituvchi yoki superuser sertifikat yuklashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        user = self.request.user
        student = serializer.validated_data.get("student")
        enrollment = student.enrollments.filter(is_active=True).first() if student else None

        if user.role == User.Role.COORDINATOR:
            coordinator = user
        else:
            # Uploader isn't the coordinator themself (e.g. superuser) —
            # fall back to the student's currently assigned coordinator, if any.
            coordinator = enrollment.coordinator if (enrollment and enrollment.coordinator) else None
        cert = serializer.save(coordinator=coordinator, branch_id=getattr(user, "branch_id", None), is_active=True)

        student_label = cert.student.full_name or cert.student.phone
        coordinator_label = cert.coordinator.full_name if cert.coordinator else (user.full_name or user.phone)

        # No payment is created here — a superuser creates the real bonus
        # payment (with an actual amount) later via pay_bonus, once the cert
        # is reviewed. See StudentCertificateViewSet.pay_bonus.

        Notification.objects.create(
            title=f"Yangi sertifikat yuklandi: {student_label}",
            note=f"Yuklagan: {coordinator_label}",
            status=Notification.Status.CERTIFICATE_UPLOAD,
            target_id=cert.student_id,
            user=None,
            branch=None,
        )

    # create() above restricts who may upload a certificate — matches
    # canUploadCertificates (admin/superuser/instructor/coordinator) for
    # editing; destroy (which would erase exam-pass proof, and is
    # irreversible for the student's own record-keeping) is tightened
    # further to admin/superuser only. Both were previously unrestricted.
    def update(self, request, *args, **kwargs):
        user = request.user
        if not (is_admin_or_superuser(user) or (user and user.is_authenticated and user.role in (User.Role.INSTRUCTOR, User.Role.COORDINATOR))):
            return Response({"detail": "Sertifikatni faqat o'qituvchi, instruktor yoki admin/superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        user = request.user
        if not (is_admin_or_superuser(user) or (user and user.is_authenticated and user.role in (User.Role.INSTRUCTOR, User.Role.COORDINATOR))):
            return Response({"detail": "Sertifikatni faqat o'qituvchi, instruktor yoki admin/superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Sertifikatni faqat admin yoki superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["post"], url_path="pay-bonus")
    def pay_bonus(self, request, pk=None):
        """
        Creates the bonus_teacher payment for this exam-pass certificate at
        the given amount/method, attributed to the enrollment's coordinator
        (the teacher). No payment exists until this is called — superuser only.
        """
        cert = self.get_object()
        if not (request.user and (request.user.is_superuser or request.user.role == User.Role.SUPERUSER)):
            return Response(
                {"detail": "Faqat superuser bonus to'lashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )
        if cert.bonus_payment_id:
            return Response(
                {"detail": "Ushbu sertifikat uchun bonus allaqachon to'langan."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            amount = int(request.data.get("amount"))
            if amount <= 0:
                raise ValueError
        except (TypeError, ValueError):
            return Response({"detail": "To'g'ri summa kiriting."}, status=status.HTTP_400_BAD_REQUEST)

        method = request.data.get("method") or Payment.Method.CASH
        enrollment = cert.student.enrollments.filter(is_active=True).first() if cert.student_id else None
        coordinator = enrollment.coordinator if (enrollment and enrollment.coordinator) else None
        if not coordinator:
            return Response(
                {"detail": "Ushbu o'quvchi uchun biriktirilgan o'qituvchi topilmadi."},
                status=status.HTTP_400_BAD_REQUEST
            )

        student_label = cert.student.full_name or cert.student.phone
        payment = Payment.objects.create(
            user=coordinator,
            enrollment=enrollment,
            amount=amount,
            status=Payment.Status.BONUS_TEACHER,
            method=method,
            branch=cert.branch,
            notes=f"Sertifikat bonusi: {student_label}",
        )
        cert.bonus_payment = payment
        cert.save(update_fields=["bonus_payment", "updated_at"])

        return Response(StudentCertificateSerializer(cert, context={"request": request}).data)


# ---------------------------------------------------------------------------
# Attendance
# ---------------------------------------------------------------------------

class AttendanceViewSet(SoftDeleteModelViewSet):
    """
    Per-day absence marks ("Yo'qlama") a teacher records for their assigned
    students. Only today's date may ever be written — past days are frozen
    history, future days don't exist yet. `create` doubles as the toggle
    endpoint: posting the same (enrollment, date) again updates the existing
    mark instead of erroring on the unique constraint.
    """

    queryset = Attendance.objects.filter(is_active=True).order_by("-date")
    serializer_class = AttendanceSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        enrollment = self.request.query_params.get("enrollment")
        enrollment_in = self.request.query_params.get("enrollment__in")
        student = self.request.query_params.get("student")
        date_param = self.request.query_params.get("date")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")

        if enrollment:
            qs = qs.filter(enrollment_id=enrollment)
        if enrollment_in:
            ids = [v for v in enrollment_in.split(",") if v.strip().isdigit()]
            qs = qs.filter(enrollment_id__in=ids)
        if student:
            qs = qs.filter(enrollment__student_id=student)
        if date_param:
            qs = qs.filter(date=date_param)
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        return qs

    @action(detail=False, methods=["get"], url_path="history")
    def history(self, request):
        """
        Per-enrollment, day-by-day attendance status from the student's own
        date_joined through today — every calendar day, not just the
        group's scheduled lesson days. This intentionally does not look at
        learning_days or the group's started_at/ends_at at all; it only
        needs the student's own join date and whatever Attendance rows
        already exist for the enrollment.

        Response: {enrollment_id: [{"date": "YYYY-MM-DD", "status": "keldi"|"kelmadi"|"today"|"unknown"}, ...]}
        Each enrollment's list is newest-first (today first).
        """
        enrollment_in = request.query_params.get("enrollment__in")
        enrollment_param = request.query_params.get("enrollment")
        ids = []
        if enrollment_in:
            ids = [int(v) for v in enrollment_in.split(",") if v.strip().isdigit()]
        elif enrollment_param and str(enrollment_param).isdigit():
            ids = [int(enrollment_param)]
        if not ids:
            return Response({})

        enrollments = Enrollment.objects.filter(pk__in=ids, is_active=True).select_related("student").only(
            "id", "created_at", "student__id", "student__date_joined"
        )
        records = Attendance.objects.filter(enrollment_id__in=ids, is_active=True).values("enrollment_id", "date", "is_absent")

        marks_by_enrollment = {}
        for r in records:
            marks_by_enrollment.setdefault(r["enrollment_id"], {})[r["date"].isoformat()] = r["is_absent"]

        today = timezone.localdate()
        result = {}
        for e in enrollments:
            joined = e.student.date_joined if e.student and e.student.date_joined else None
            start = joined.date() if joined else (e.created_at.date() if e.created_at else today)
            marks = marks_by_enrollment.get(e.id, {})
            days = []
            cursor = today
            while cursor >= start:
                date_str = cursor.isoformat()
                is_absent = marks.get(date_str)
                if is_absent is None:
                    day_status = "today" if cursor == today else "unknown"
                else:
                    day_status = "kelmadi" if is_absent else "keldi"
                days.append({"date": date_str, "status": day_status})
                cursor -= timedelta(days=1)
            result[str(e.id)] = days

        return Response(result)

    @staticmethod
    def _can_mark(user, enrollment):
        if is_admin_or_superuser(user):
            return True
        if not (user and user.is_authenticated):
            return False
        if user.role == User.Role.COORDINATOR and enrollment.coordinator_id == user.id:
            return True
        if user.role == User.Role.INSTRUCTOR and enrollment.instructor_id == user.id:
            return True
        return False

    def create(self, request, *args, **kwargs):
        enrollment_id = request.data.get("enrollment")
        try:
            enrollment = Enrollment.objects.get(pk=enrollment_id, is_active=True)
        except (Enrollment.DoesNotExist, TypeError, ValueError):
            return Response({"detail": "O'quvchi biriktiruvi topilmadi."}, status=status.HTTP_400_BAD_REQUEST)

        if not self._can_mark(request.user, enrollment):
            return Response(
                {"detail": "Faqat ushbu o'quvchiga biriktirilgan o'qituvchi/instruktor yoki admin davomatni belgilashi mumkin."},
                status=status.HTTP_403_FORBIDDEN
            )

        if enrollment.status in (Enrollment.Status.FINISHED, Enrollment.Status.CANCELED):
            return Response(
                {"detail": "O'quvchi tugatgan yoki bekor qilingan bo'lsa, davomat belgilab bo'lmaydi."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if enrollment.group_id and enrollment.group.status in (Group.Status.FINISHED, Group.Status.CANCELED):
            return Response(
                {"detail": "Guruh tugatgan yoki bekor qilingan bo'lsa, davomat belgilab bo'lmaydi."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # The date is always the server's current date — a client-supplied
        # date is only accepted if it matches today, so a teacher can never
        # backdate or forward-date an absence mark.
        today = timezone.localdate()
        date_val = request.data.get("date") or str(today)
        if str(date_val) != str(today):
            return Response(
                {"detail": "Faqat bugungi kun uchun davomat belgilash mumkin."},
                status=status.HTTP_400_BAD_REQUEST
            )

        is_absent = bool(request.data.get("is_absent"))
        record, _ = Attendance.objects.update_or_create(
            enrollment=enrollment,
            date=today,
            defaults={"is_absent": is_absent, "marked_by": request.user, "is_active": True},
        )
        return Response(AttendanceSerializer(record).data, status=status.HTTP_200_OK)

    # create() above (which doubles as the toggle) enforces _can_mark's
    # ownership check plus the "only today, never backdated" rule — but
    # the standard update/partial_update/destroy endpoints had none of
    # that, so *any* authenticated user could PATCH an arbitrary
    # historical Attendance row's is_absent (rewriting the audit trail
    # for a student they have no connection to) or DELETE one outright.
    # Restricted to admin/superuser, who may need to correct a mistaken
    # mark after the fact.
    def update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Davomat yozuvini faqat admin yoki superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Davomat yozuvini faqat admin yoki superuser tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin_or_superuser(request.user):
            return Response({"detail": "Davomat yozuvini faqat admin yoki superuser o'chirishi mumkin."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Legacy-data Excel import (superuser only) — see management/import_excel.py
# ---------------------------------------------------------------------------

@api_view(["POST"])
def import_excel_upload(request):
    """
    Accepts a multipart-uploaded Excel workbook, stashes it under
    MEDIA_ROOT/imports/, and hands it off to a Celery task — the file has
    thousands of rows, far past what should run inline in a request/response
    cycle. Returns a task id the frontend polls via import_excel_status.
    """
    if not is_superuser(request.user):
        return Response(
            {"detail": "Excel orqali import qilish faqat superuser uchun ruxsat etilgan."},
            status=status.HTTP_403_FORBIDDEN,
        )

    uploaded = request.FILES.get("file")
    if not uploaded:
        return Response({"detail": "Fayl yuborilmadi."}, status=status.HTTP_400_BAD_REQUEST)

    allowed_extensions = (".xlsx", ".xlsm", ".xls")
    if not uploaded.name.lower().endswith(allowed_extensions):
        return Response(
            {"detail": "Faqat Excel fayl (.xlsx, .xlsm, .xls) yuklash mumkin."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    from management.tasks import import_excel_task

    imports_dir = os.path.join(settings.MEDIA_ROOT, "imports")
    os.makedirs(imports_dir, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex}_{os.path.basename(uploaded.name)}"
    saved_path = os.path.join(imports_dir, safe_name)
    with open(saved_path, "wb") as f:
        for chunk in uploaded.chunks():
            f.write(chunk)

    task = import_excel_task.delay(saved_path, request.user.id)
    return Response({"task_id": task.id}, status=status.HTTP_202_ACCEPTED)


@api_view(["GET"])
def import_excel_status(request, task_id):
    """Polled by the frontend while the import task runs in the background."""
    if not is_superuser(request.user):
        return Response(
            {"detail": "Excel orqali import qilish faqat superuser uchun ruxsat etilgan."},
            status=status.HTTP_403_FORBIDDEN,
        )

    result = AsyncResult(task_id)
    payload = {"state": result.state}
    if result.state == "SUCCESS":
        payload["result"] = result.result
    elif result.state == "FAILURE":
        payload["error"] = str(result.result)
    return Response(payload)


