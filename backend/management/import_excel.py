"""
management/import_excel.py

One-off bulk importer for the legacy "Тўлов" Excel workbook (groups,
students, their accepted-payment installments, and agent referral
bonuses), read from the "Гурух" sheet only — the "Касса" sheet is a
separate cash-log and is intentionally not touched.

Column layout of the "Гурух" sheet (0-indexed, as read by openpyxl):

    C(2)  group name            D(3)  category code
    F(5)  group start date      G(6)  group end date
    J(9)  student full name     K(10) student JSHSHR
    L(11) passport "AB 1234567" M(12) birth date
    N(13) phone                 O(14) phone2
    P(15) referring agent's name/nickname
    Q..U(16-20)                 5 accepted-payment installments (so'm x1000)
    V(21) agent bonus amount (so'm x1000)   W(22) bonus payment date
    Y(24) course-completion certificate     Z(25) enrollment note

Everything else (A, B, E, H, I, X) is intentionally ignored — see the
column-mapping conversation in the PR/commit that added this file for why.

Entry point: `import_excel_data(file_path, actor, branch)`. Runs inside a single
`transaction.atomic()` block — either the whole file lands, or none of it
does (an unhandled error rolls everything back). Individually messy rows
(a garbled certificate number, an unparsable passport, ...) do NOT abort
the import; they're skipped/degraded gracefully and recorded in the
returned summary's "warnings" list instead.
"""

from datetime import date, datetime, time

import openpyxl
from django.db import transaction
from django.utils import timezone

from management.helpers import (
    ImportContext,
    clean_text,
    get_or_create_agent,
    get_or_create_category,
    get_or_create_group,
    get_or_create_student,
    parse_certificate,
    parse_excel_date,
    parse_full_name,
    parse_jshshr,
    parse_passport,
    normalize_phone,
    to_money,
)
from management.models import Enrollment, Group, Payment

SHEET_NAME = "Гурух"

COL_GROUP = 2
COL_CATEGORY = 3
COL_GROUP_START = 5
COL_GROUP_END = 6
COL_FULL_NAME = 9
COL_JSHSHR = 10
COL_PASSPORT = 11
COL_BIRTH_DATE = 12
COL_PHONE = 13
COL_PHONE2 = 14
COL_AGENT = 15
COLS_PAYMENTS = range(16, 21)  # Q..U
COL_BONUS_AMOUNT = 21
COL_BONUS_DATE = 22
COL_CERTIFICATE = 24
COL_NOTE = 25
LAST_COL = 26  # A..Z — the sheet's nominal used-range runs far past this
                # (stray formatting inflates it to column XFD), so iter_rows
                # is bounded explicitly below or it reads ~16k cells/row.


def _payment_backdate(group):
    """
    None of the per-installment payments carry their own date in the sheet,
    only the group's start date — used as a stand-in so five-year-old
    payments don't land in "today's income" on the finance dashboard.
    Falls back to now() for a group with no start date, or with a start
    date in the future (a not-yet-started group can't have payments dated
    on a day that hasn't happened yet — that would otherwise show up as
    future-dated "accepted" income).
    """
    today = timezone.localdate()
    if group.started_at and group.started_at <= today:
        return timezone.make_aware(datetime.combine(group.started_at, time(hour=12)))
    return timezone.now()


def _import_row(ctx, row_number, row, actor, branch):
    group_name = row[COL_GROUP]
    category_code = row[COL_CATEGORY]
    full_name_raw = row[COL_FULL_NAME]

    if group_name is None and full_name_raw is None:
        return  # fully blank row

    group_name = str(group_name).strip() if group_name is not None else None
    if not group_name:
        ctx.warn(row_number, "guruh nomi yo'q, qator o'tkazib yuborildi.")
        return

    category_name = str(category_code).strip() if category_code else None
    if not category_name:
        ctx.warn(row_number, f"'{group_name}' guruhi uchun kategoriya yo'q, qator o'tkazib yuborildi.")
        return
    category = get_or_create_category(ctx, category_name)

    started_at = parse_excel_date(row[COL_GROUP_START])
    ends_at = parse_excel_date(row[COL_GROUP_END])
    group = get_or_create_group(ctx, group_name, category, started_at, ends_at, branch)

    full_name = parse_full_name(full_name_raw)
    if not full_name:
        # A group-shell row with no student on it yet (e.g. a future group
        # pre-created with placeholder rows) — the group above is still
        # created/looked up, there's just no one to enroll.
        return

    jshshr = parse_jshshr(row[COL_JSHSHR])
    phone = normalize_phone(row[COL_PHONE])
    phone2 = clean_text(row[COL_PHONE2])
    passport_serie, passport_number = parse_passport(row[COL_PASSPORT])
    birth_date = parse_excel_date(row[COL_BIRTH_DATE])

    student, _created = get_or_create_student(
        ctx, row_number, jshshr, phone, full_name, phone2, passport_serie, passport_number, birth_date, branch,
    )

    cert_series, cert_number, cert_leftover = parse_certificate(row[COL_CERTIFICATE])
    if cert_number and not student.certificate_number:
        student.certificate_series = cert_series
        student.certificate_number = cert_number
        student.save(update_fields=["certificate_series", "certificate_number", "updated_at"])
    elif cert_leftover:
        ctx.warn(row_number, f"guvohnoma ustunidagi qiymat tushunarsiz: '{cert_leftover}' — izohga qo'shildi.")

    if Enrollment.objects.filter(student=student, group=group).exists():
        ctx.warn(row_number, f"{full_name} allaqachon '{group_name}' guruhida ro'yxatdan o'tgan, qator o'tkazib yuborildi.")
        return

    note_parts = []
    raw_note = row[COL_NOTE]
    if raw_note:
        note_parts.append(str(raw_note).strip())
    if cert_leftover:
        note_parts.append(f"Guvohnoma ustuni: {cert_leftover}")

    status = Enrollment.Status.FINISHED if group.status == Group.Status.FINISHED else Enrollment.Status.ENROLLED

    agent = None
    agent_raw = row[COL_AGENT]
    if agent_raw and str(agent_raw).strip():
        agent = get_or_create_agent(ctx, str(agent_raw).strip(), branch)

    enrollment = Enrollment.objects.create(
        branch=branch,
        student=student,
        category=category,
        group=group,
        agent=agent,
        status=status,
        notes="; ".join(note_parts) or None,
        excel_imported=True,
    )
    ctx.created["enrollments"] += 1

    payment_date = _payment_backdate(group)
    new_payments = []
    installment_count = 0
    for col in COLS_PAYMENTS:
        amount = to_money(row[col])
        if amount <= 0:
            continue
        new_payments.append(Payment(
            branch=branch,
            user=actor,
            created_by=actor,
            enrollment=enrollment,
            amount=amount,
            status=Payment.Status.ACCEPTED,
            created_at=payment_date,
        ))
        installment_count += 1

    bonus_amount = to_money(row[COL_BONUS_AMOUNT])
    bonus_added = False
    if bonus_amount > 0 and agent is not None:
        bonus_date = parse_excel_date(row[COL_BONUS_DATE])
        bonus_datetime = (
            timezone.make_aware(datetime.combine(bonus_date, time(hour=12))) if bonus_date else payment_date
        )
        new_payments.append(Payment(
            branch=branch,
            user=actor,
            created_by=actor,
            enrollment=enrollment,
            agent=agent,
            amount=bonus_amount,
            status=Payment.Status.BONUS,
            created_at=bonus_datetime,
            notes=f"Agentlik bonusi: {full_name}",
        ))
        bonus_added = True
    elif bonus_amount > 0 and agent is None:
        ctx.warn(row_number, f"{full_name} uchun bonus summasi bor ({bonus_amount}) lekin agent ko'rsatilmagan.")

    if new_payments:
        Payment.objects.bulk_create(new_payments)
        ctx.created["payments"] += installment_count
        if bonus_added:
            ctx.created["bonus_payments"] += 1


def import_excel_data(file_path, actor, branch):
    """
    Runs the full import inside one atomic transaction and returns a
    summary dict: {"created": {...counts...}, "warnings": [...]}.

    `branch` is required — every group/student/enrollment/payment/agent
    created (or matched) by this run is tagged with it, since the legacy
    workbook itself carries no branch information of its own.
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Faylda '{SHEET_NAME}' varag'i topilmadi.")
    sheet = workbook[SHEET_NAME]

    ctx = ImportContext()
    with transaction.atomic():
        rows = sheet.iter_rows(min_row=2, max_col=LAST_COL, values_only=True)
        for row_number, row in enumerate(rows, start=2):
            _import_row(ctx, row_number, row, actor, branch)

    return {"created": ctx.created, "warnings": ctx.warnings}
